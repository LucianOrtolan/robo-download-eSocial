import time
from datetime import datetime, timedelta
import PySimpleGUI as sg
import openpyxl
from openpyxl.utils.cell import get_column_letter
import undetected_chromedriver as uc
from auto_download_undetected_chromedriver import download_undetected_chromedriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from undetected_chromedriver import ChromeOptions
import os
import shutil
import requests

# Layout
sg.theme('Dark Blue 3')

layout = [[sg.Text('Caminho da planilha: ', size=20), sg.InputText('', key='caminho'),
sg.FileBrowse("Procurar", file_types=(("Excel", "*.xlsx"),))],
[sg.Text('Linha inicial da planilha: ', size=20), sg.InputText("1", key='linhaini', size=4)],
[sg.Text('Linha final da planilha: ', size=20), sg.InputText("1", key='linhafim', size=4)],
[sg.Text('Meses a buscar: ', size=20), sg.DropDown(["1", "2", "3", "4", "5", "6"], "3", key='periodo')],
[sg.Text('Solicitar (1) / Baixar (2): ', size=20), sg.DropDown(["1", "2"], "1", key='finalidade')],
[sg.Text('Salvar arquivos: ', size=20), sg.InputText('', key='salvar'), sg.FolderBrowse('Procurar')],
[sg.Text('Data inicial (DD/MM/AAAA): ', size=20), sg.InputText('', size=10, key='dataini')],
[sg.Checkbox('Certificado Próprio', key='proprio')],
[sg.Text('1. Se selecionada a opção "Certificado Próprio" a planilha não precisa ser informada')],
[sg.Button('Iniciar'), sg.Button('Sair')]]

# Janela
janela = sg.Window('Download eSocial', layout)

while True:
    eventos, valores = janela.read()                

    if eventos == sg.WIN_CLOSED or eventos == 'Sair':
        break
    if eventos == 'Iniciar':

        download_dir = valores['salvar'].rstrip().replace('/', '\\')

        def criar_pasta(nome_empresa):
            pasta_empresa = os.path.join(download_dir, nome_empresa)
            if not os.path.exists(pasta_empresa):
                os.makedirs(pasta_empresa)
            return pasta_empresa

        url = 'https://login.esocial.gov.br/login.aspx'
        folder_path = "c:\\chromedriver"

        if valores['proprio'] is False:
            workbook = openpyxl.load_workbook(valores['caminho'])
            sheet_empresas = workbook.active
            documento = sheet_empresas.cell(row=int(valores['linhaini']), column=3).value                    

        # Condição para solicitar as datas quando o certificado é por procuração de PJ/PF
        if valores['finalidade'] == '1' and valores['proprio'] is False:
            chrome_options = ChromeOptions()
            chromedriver_path = download_undetected_chromedriver(folder_path, undetected=True, arm=False,
                                                                force_update=True)
            driver = uc.Chrome(driver_executable_path=chromedriver_path, options=chrome_options, headless=False)            
            driver.get(url)
            driver.maximize_window()                                                
            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="login-acoes"]/div[2]/p/button'))
            )
            driver.find_element('xpath', '//*[@id="login-acoes"]/div[2]/p/button').click()

            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="login-certificate"]'))
            )
            driver.find_element('xpath', '//*[@id="login-certificate"]').click()

            print("Selecione o certificado para continuar")

            janela.refresh()
            janela.minimize()

            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="sairAplicacao"]'))
            )
            
            inscricao = driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/p[2]/span[1]').text.strip('-')
            print(f'CNPJ do procurador: {inscricao}')
            # Condição para identificar se a inscrição é um CNPJ ou CPF
            if len(driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/p[2]/span[1]').text) < 18:
                driver.find_element(By.XPATH, '//*[@id="geral"]/div').click()
                WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="header"]/div[2]/a'))
                )
                driver.find_element('xpath', '//*[@id="header"]/div[2]/a').click()
                pass
            else:
                driver.find_element(By.CLASS_NAME, 'alterar-perfil').click()

            empresas = list(sheet_empresas.iter_rows(min_row=int(valores['linhaini']), max_row=int(valores['linhafim'])))                      

            if valores['dataini']:
                data_inicial = datetime.strptime(valores['dataini'], '%d/%m/%Y')
            else:
                data_inicial = datetime(2018, 1, 1)
            
            # Dicionário para armazenar a última data final de cada empresa
            ultima_data_final_por_empresa = {}    

            data_corte = ''                   
            loop = True
            while loop:
                for linha in empresas:
                    documento = len(str(linha[2].value))
                    data_atual = data_inicial
                    # Condição que verifica se é CNPJ ou CPF na planilha                                
                    if documento >= 15:
                        cnpj = linha[2].value # CNPJ
                        driver.find_element('xpath', '//*[@id="perfilAcesso"]').click()
                        driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.DOWN + Keys.DOWN)
                        driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.ENTER)
                        driver.find_element('xpath', '//*[@id="procuradorCnpj"]').send_keys(cnpj)
                        driver.find_element('xpath', '//*[@id="procuradorCnpj"]').send_keys(Keys.LEFT_CONTROL + 'v')
                        WebDriverWait(driver, 120).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@id="btn-verificar-procuracao-cnpj"]'))
                            )
                        driver.find_element('xpath', '//*[@id="btn-verificar-procuracao-cnpj"]').click()
                        mensagem_procuracao = ''
                        try:
                            WebDriverWait(driver, 15).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                            )
                        except:
                            mensagem_procuracao = driver.find_element(By.CLASS_NAME, 'fade-alert').text[2:]

                        # Condição se verifica se possui procuração para o CNPJ que está sendo procurado
                        print(mensagem_procuracao)
                        if mensagem_procuracao == 'O procurador não possui perfil com autorização de acesso à Web':
                            print(f'Não possui procuração para o {cnpj}')
                            linha_celula = linha[4]

                            if hasattr(linha_celula, 'row'):
                                linha_atual = linha_celula.row
                                sheet_empresas[f'F{linha_atual}'] = 'Não possui procuração'
                                workbook.save(valores['caminho'])
                                print('Retornando as buscas')
                                driver.refresh()
                                continue
                        else:
                            print(f'CNPJ/CPF sendo buscado: {str(linha[2].value)}')
                            driver.find_element('xpath', '//*[@id="geral"]/div').click()
                            WebDriverWait(driver, 120).until(
                                EC.presence_of_element_located((By.XPATH, '//*[@id="menuDownload"]'))
                            )
                            driver.find_element('xpath', '//*[@id="menuDownload"]').click()
                            driver.find_element('xpath', '//*[@id="menuDownload"]').send_keys(Keys.DOWN + Keys.ENTER)
                            WebDriverWait(driver, 120).until(
                                EC.presence_of_element_located((By.XPATH, '//*[@id="TipoPedido"]'))
                            )
                                
                            data_corte = datetime.strptime(driver.find_element(By.CLASS_NAME, 'alert-info').text[58:68],'%d/%m/%Y')    

                            # Obtém a última data final para a empresa
                            if linha not in ultima_data_final_por_empresa:
                                data_atual = data_inicial
                                ultima_data_final_por_empresa[linha] = data_atual
                            data_final = ultima_data_final_por_empresa[linha] + timedelta(days=30 * int(valores['periodo']))

                            # Fazer cinco requisições para a empresa atual
                            for i in range(5):
                                if data_final > data_corte:
                                    data_final = data_corte
                                    loop = False

                                data_inicial_str = ultima_data_final_por_empresa[linha].strftime('%d/%m/%Y')
                                data_final_str = data_final.strftime('%d/%m/%Y')

                                driver.find_element('xpath', '//*[@id="TipoPedido"]/option[2]').click()
                                driver.find_element('xpath', '//*[@id="DataInicial"]').send_keys(data_inicial_str)
                                driver.find_element('xpath', '//*[@id="DataFinal"]').click()
                                driver.find_element('xpath', '//*[@id="DataFinal"]').clear()
                                driver.find_element('xpath', '//*[@id="DataFinal"]').send_keys(data_final_str)
                                print(f'Data inicial: {data_inicial_str} - Data Final: {data_final_str}')
                                driver.find_element('xpath', '//*[@id="btnSalvar"]').click()
                                pedido = driver.find_element(By.CLASS_NAME, 'fade-alert').text[2:]
                                print(pedido)
                                if pedido == 'Solicitação enviada com sucesso.':
                                    WebDriverWait(driver, 120).until(
                                        EC.presence_of_element_located((By.XPATH, '//*[@id="conteudo-pagina"]/div[1]/a'))
                                    )
                                    driver.find_element('xpath', '//*[@id="conteudo-pagina"]/div[1]/a').click()
                                    ultima_data_final_por_empresa[linha] = data_final + timedelta(days=1)
                                    data_final = ultima_data_final_por_empresa[linha] + timedelta(days=30 * int(valores['periodo']))

                                    linha_celula = linha[4]     
                                    if hasattr(linha_celula, 'row'):
                                        linha_atual = linha_celula.row
                                        sheet_empresas[f'F{linha_atual}'] = 'Solicitação realizada'
                                        sheet_empresas[f"G{linha_atual}"] = "Última data solicitada: " + data_final_str
                                        workbook.save(valores['caminho'])

                                elif pedido == 'Pedido não foi aceito. Já existe um pedido do mesmo tipo.':
                                    driver.find_element(By.XPATH, '//*[@id="btnCancelarAlteracao"]').click()
                                    WebDriverWait(driver, 120).until(
                                        EC.presence_of_element_located((By.XPATH, '//*[@id="conteudo-pagina"]/div[1]/a'))
                                    )
                                    driver.find_element('xpath', '//*[@id="conteudo-pagina"]/div[1]/a').click()
                                    ultima_data_final_por_empresa[linha] = data_final + timedelta(days=1)
                                    data_final = ultima_data_final_por_empresa[linha] + timedelta(days=30 * int(valores['periodo']))                                               
                                                                                    

                                elif pedido == 'O limite de solicitações foi alcançado. Somente é permitido 72 (doze) solicitações por dia.':
                                    time.sleep(2)
                                    driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/a').click()
                                    break
                            
                        # Atualiza a data atual para a próxima iteração
                        data_atual = ultima_data_final_por_empresa[empresas[0]] + timedelta(days=1)
                        driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/a').click()                                

                    else:
                        # Buscas por CPF
                        cnpj = linha[2].value
                        driver.find_element('xpath', '//*[@id="perfilAcesso"]').click()
                        driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.DOWN)
                        driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.ENTER)
                        driver.find_element('xpath', '//*[@id="procuradorCpf"]').send_keys(cnpj)
                        driver.find_element('xpath', '//*[@id="procuradorCpf"]').send_keys(Keys.LEFT_CONTROL + 'v')
                        WebDriverWait(driver, 120).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@id="btn-verificar-procuracao-cpf"]'))
                            )
                        driver.find_element('xpath', '//*[@id="btn-verificar-procuracao-cpf"]').click()
                        mensagem_procuracao = ''
                        try:
                            WebDriverWait(driver, 15).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                            )
                        except:
                            mensagem_procuracao = driver.find_element(By.CLASS_NAME, 'fade-alert').text[2:]
                        
                        # Condição se verifica se possui procuração para o CNPJ que está sendo procurado
                        print(mensagem_procuracao)
                        if mensagem_procuracao == 'O procurador não possui perfil com autorização de acesso à Web':
                            print(f'Não possui procuração para o {cnpj}')
                            linha_celula = linha[4]

                            if hasattr(linha_celula, 'row'):
                                linha_atual = linha_celula.row
                                sheet_empresas[f'F{linha_atual}'] = 'Não possui procuração'
                                workbook.save(valores['caminho'])
                                print('Retornando as buscas')
                                driver.refresh()
                                pass
                        else:
                            print(f'CNPJ/CPF sendo buscado: {str(linha[2].value)}')
                            driver.find_element('xpath', '//*[@id="geral"]/div').click()
                            WebDriverWait(driver, 120).until(
                                EC.presence_of_element_located((By.XPATH, '//*[@id="menuDownload"]'))
                            )
                            driver.find_element('xpath', '//*[@id="menuDownload"]').click()
                            driver.find_element('xpath', '//*[@id="menuDownload"]').send_keys(Keys.DOWN + Keys.ENTER)
                            WebDriverWait(driver, 120).until(
                                EC.presence_of_element_located((By.XPATH, '//*[@id="TipoPedido"]'))
                            )

                            data_corte = datetime.strptime(driver.find_element(By.CLASS_NAME, 'alert-info').text[58:68],'%d/%m/%Y')    

                            # Obtém a última data final para a empresa
                            if linha not in ultima_data_final_por_empresa:
                                data_atual = data_inicial
                                ultima_data_final_por_empresa[linha] = data_atual
                            data_final = ultima_data_final_por_empresa[linha] + timedelta(days=30)

                            # Fazer cinco requisições para a empresa atual
                            for i in range(5):
                                if data_final > data_corte:
                                    data_final = data_corte
                                    break

                                data_inicial_str = ultima_data_final_por_empresa[linha].strftime('%d/%m/%Y')
                                data_final_str = data_final.strftime('%d/%m/%Y')

                                driver.find_element('xpath', '//*[@id="TipoPedido"]/option[2]').click()
                                driver.find_element('xpath', '//*[@id="DataInicial"]').send_keys(data_inicial_str)
                                driver.find_element('xpath', '//*[@id="DataFinal"]').click()
                                driver.find_element('xpath', '//*[@id="DataFinal"]').clear()
                                driver.find_element('xpath', '//*[@id="DataFinal"]').send_keys(data_final_str)
                                print(f'Data inicial: {data_inicial_str} - Data Final: {data_final_str}')
                                driver.find_element('xpath', '//*[@id="btnSalvar"]').click()
                                pedido = driver.find_element(By.CLASS_NAME, 'fade-alert').text[2:]
                                print(pedido)
                                if pedido == 'Solicitação enviada com sucesso.':
                                    WebDriverWait(driver, 120).until(
                                        EC.presence_of_element_located((By.XPATH, '//*[@id="conteudo-pagina"]/div[1]/a'))
                                    )
                                    driver.find_element('xpath', '//*[@id="conteudo-pagina"]/div[1]/a').click()
                                    ultima_data_final_por_empresa[linha] = data_final + timedelta(days=1)
                                    data_final = ultima_data_final_por_empresa[linha] + timedelta(days=30 * int(valores['periodo']))

                                    linha_celula = linha[4]
                                    if hasattr(linha_celula, 'row'):
                                        linha_atual = linha_celula.row
                                        sheet_empresas[f'F{linha_atual}'] = 'Solicitação realizada'
                                        sheet_empresas[f"G{linha_atual}"] = "Última data solicitada: " + data_final_str
                                        workbook.save(valores['caminho'])

                                elif pedido == 'Pedido não foi aceito. Já existe um pedido do mesmo tipo.':
                                    driver.find_element(By.XPATH, '//*[@id="btnCancelarAlteracao"]').click()
                                    WebDriverWait(driver, 120).until(
                                        EC.presence_of_element_located((By.XPATH, '//*[@id="conteudo-pagina"]/div[1]/a'))
                                    )
                                    driver.find_element('xpath', '//*[@id="conteudo-pagina"]/div[1]/a').click()
                                    ultima_data_final_por_empresa[linha] = data_final + timedelta(days=1)
                                    data_final = ultima_data_final_por_empresa[linha] + timedelta(days=30 * int(valores['periodo']))

                                elif pedido == 'O limite de solicitações foi alcançado. Somente é permitido 72 (doze) solicitações por dia.':
                                    time.sleep(2)
                                    driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/a').click()
                                    break
                            
                        # Atualiza a data atual para a próxima iteração
                        data_atual = ultima_data_final_por_empresa[empresas[0]] + timedelta(days=1)
                        driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/a').click()
            
            print('Buscas Finalizadas')
            time.sleep(7)
            driver.quit()

        if valores['finalidade'] == '2' and valores['proprio'] is False:
            chrome_options = ChromeOptions()
            prefs = {'download.default_directory': download_dir}
            chrome_options.add_experimental_option('prefs', prefs)
            chromedriver_path = download_undetected_chromedriver(folder_path, undetected=True, arm=False,
                                                                force_update=True)
            driver = uc.Chrome(driver_executable_path=chromedriver_path, options=chrome_options,headless=False)            
            driver.get(url)
            driver.maximize_window()
            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="login-acoes"]/div[2]/p/button'))
            )
            driver.find_element('xpath', '//*[@id="login-acoes"]/div[2]/p/button').click()

            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="login-certificate"]'))
            )
            driver.find_element('xpath', '//*[@id="login-certificate"]').click()

            print("Selecione o certificado para continuar")

            janela.refresh()
            janela.minimize()

            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="sairAplicacao"]'))
            )            
            inscricao = driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/p[2]/span[1]').text.strip('-')
            print(f'CNPJ do procurador: {inscricao}')
            if len(driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/p[2]/span[1]').text) < 18:
                driver.find_element(By.XPATH, '//*[@id="geral"]/div').click()
                WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="header"]/div[2]/a'))
                )
                driver.find_element('xpath', '//*[@id="header"]/div[2]/a').click()
                pass
            else:
                driver.find_element(By.CLASS_NAME, 'alterar-perfil').click()

            for linha in sheet_empresas.iter_rows(min_row=int(valores['linhaini']), max_row=int(valores['linhafim'])):
                pasta_empresa = criar_pasta(f'{linha[0].value} - {linha[1].value.rstrip()}')
                documento = len(str(linha[2].value))
                if documento >= 15: #CNPJ
                    cnpj = linha[2].value
                    driver.find_element('xpath', '//*[@id="perfilAcesso"]').click()
                    driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.DOWN + Keys.DOWN)
                    driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.ENTER)
                    driver.find_element('xpath', '//*[@id="procuradorCnpj"]').send_keys(cnpj)
                    driver.find_element('xpath', '//*[@id="procuradorCnpj"]').send_keys(Keys.LEFT_CONTROL + 'v')
                    WebDriverWait(driver, 120).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@id="btn-verificar-procuracao-cnpj"]'))
                            )
                    driver.find_element('xpath', '//*[@id="btn-verificar-procuracao-cnpj"]').click()
                    mensagem_procuracao = ''
                    try:
                        WebDriverWait(driver, 15).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                        )
                    except:
                        mensagem_procuracao = driver.find_element(By.CLASS_NAME, 'fade-alert').text[2:]

                    # Condição se verifica se possui procuração para o CNPJ que está sendo procurado                                
                    if mensagem_procuracao:
                        print(f'Não possui procuração para o {cnpj}')
                        linha_celula = linha[4]

                        if hasattr(linha_celula, 'row'):
                            linha_atual = linha_celula.row
                            sheet_empresas[f'H{linha_atual}'] = 'Não possui procuração'
                            workbook.save(valores['caminho'])
                            print('Retornando as buscas')
                            driver.refresh()
                            pass
                    else:
                        WebDriverWait(driver, 120).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                        )
                        driver.find_element('xpath', '//*[@id="geral"]/div').click()
                        WebDriverWait(driver, 120).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="menuDownload"]'))
                    )
                        driver.find_element('xpath', '//*[@id="menuDownload"]').click()
                        driver.find_element('xpath', '//*[@id="menuDownload"]').send_keys(
                        Keys.DOWN + Keys.DOWN + Keys.ENTER)
                        WebDriverWait(driver, 120).until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//*[@id="conteudo-pagina"]/form/section/div/div[4]/input'))
                    )
                        driver.find_element('xpath', '//*[@id="conteudo-pagina"]/form/section/div/div[4]/input').click()
                        time.sleep(5)
                        print(f'Iniciando download da empresa CNPJ: {cnpj}')

                        download_links = driver.find_elements('xpath',
                                                                '//*[@id="DataTables_Table_0"]/tbody/tr/td/a')
                        total_files = (len(download_links))
                        print(f'Total de arquivos da primeira página: {total_files}')
                        soma_files = 0

                        for link in driver.find_elements(By.CLASS_NAME, 'icone-baixar'):
                            link.click()
                            time.sleep(7)
                            soma_files += 1
                            print(f'Baixando {soma_files}/{total_files} arquivos')

                            arquivos_baixados = os.listdir(download_dir)
                            for arquivo in arquivos_baixados:
                                if arquivo.endswith(".zip"):  # Verifica se o arquivo é um ZIP
                                    caminho_arquivo = os.path.join(download_dir, arquivo)
                                    shutil.move(caminho_arquivo, os.path.join(pasta_empresa, arquivo))
                        try:
                            WebDriverWait(driver, 10).until(
                                    EC.presence_of_element_located(
                                        (By.XPATH, '//*[@id="DataTables_Table_0_paginate"]/span/a[2]')))
                            driver.find_element('xpath', '//*[@id="DataTables_Table_0_paginate"]/span/a[2]').click()
                            download_links2 = driver.find_element('xpath', '//*[@id="DataTables_Table_0_paginate"]/span/a[2]').click()
                            total_files2 = (len(download_links2))
                            print(f'Total de arquivos da segunda página: {total_files}')
                            soma_files2 = 0
                            for link in driver.find_elements(By.CLASS_NAME, 'icone-baixar'):
                                link.click()
                                time.sleep(7)
                                soma_files2 += 1
                                print(f'Baixando {soma_files2}/{total_files2} arquivos')
                                arquivos_baixados = os.listdir(download_dir)
                                for arquivo in arquivos_baixados:
                                    if arquivo.endswith(".zip"):  # Verifica se o arquivo é um ZIP
                                        caminho_arquivo = os.path.join(download_dir, arquivo)
                                        shutil.move(caminho_arquivo, os.path.join(pasta_empresa, arquivo))

                            loop = False

                            if loop == False:
                                linha_celula = linha[4]
                                if hasattr(linha_celula, 'row'):
                                    linha_atual = linha_celula.row
                                    sheet_empresas[f'H{linha_atual}'] = str(int(soma_files)+int(soma_files2)) + "/" + str(int(total_files)+int(total_files2))
                                    workbook.save(valores['caminho'])

                            print(f'Arquivos da empresa CNPJ: {cnpj} baixados com sucesso!')
                            time.sleep(2)
                            driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/a').click()
                            time.sleep(2)
                        except:
                            print(f'Arquivos da empresa CNPJ: {cnpj} baixados com sucesso!')
                            time.sleep(2)
                            driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/a').click()
                            time.sleep(2)
                            continue

                else: #CPF
                    cnpj = linha[2].value
                    driver.find_element('xpath', '//*[@id="perfilAcesso"]').click()
                    driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.DOWN)
                    driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.ENTER)
                    driver.find_element('xpath', '//*[@id="procuradorCpf"]').send_keys(cnpj)
                    driver.find_element('xpath', '//*[@id="procuradorCpf"]').send_keys(Keys.LEFT_CONTROL + 'v')
                    WebDriverWait(driver, 120).until(
                                EC.element_to_be_clickable((By.XPATH, '//*[@id="btn-verificar-procuracao-cpf"]'))
                            )
                    driver.find_element('xpath', '//*[@id="btn-verificar-procuracao-cpf"]').click()
                    mensagem_procuracao = ''
                    try:
                        WebDriverWait(driver, 15).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                        )
                    except:
                        mensagem_procuracao = driver.find_element(By.CLASS_NAME, 'fade-alert').text[2:]

                    # Condição se verifica se possui procuração para o CNPJ que está sendo procurado
                    if mensagem_procuracao:
                        print(f'Não possui procuração para o {cnpj}')
                        linha_celula = linha[4]

                        if hasattr(linha_celula, 'row'):
                            linha_atual = linha_celula.row
                            sheet_empresas[f'H{linha_atual}'] = 'Não possui procuração'
                            workbook.save(valores['caminho'])
                            print('Retornando as buscas')
                            driver.refresh()
                            pass
                    else:
                        WebDriverWait(driver, 120).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                        )
                        driver.find_element('xpath', '//*[@id="geral"]/div').click()
                        WebDriverWait(driver, 120).until(
                            EC.presence_of_element_located((By.XPATH, '//*[@id="menuDownload"]'))
                        )
                        driver.find_element('xpath', '//*[@id="menuDownload"]').click()
                        driver.find_element('xpath', '//*[@id="menuDownload"]').send_keys(
                            Keys.DOWN + Keys.DOWN + Keys.ENTER)
                        WebDriverWait(driver, 120).until(
                            EC.presence_of_element_located(
                                (By.XPATH, '//*[@id="conteudo-pagina"]/form/section/div/div[4]/input'))
                        )
                        driver.find_element('xpath', '//*[@id="conteudo-pagina"]/form/section/div/div[4]/input').click()
                        time.sleep(5)                                    
                        print(f'"Iniciando download da empresa CNPJ: {cnpj}')

                        download_links = driver.find_elements('xpath',
                                                            '//*[@id="DataTables_Table_0"]/tbody/tr/td/a')
                        total_files = (len(download_links))
                        print(f'Total de arquivos da primeira página: {total_files}')
                        soma_files = 0

                        for link in driver.find_elements(By.CLASS_NAME, 'icone-baixar'):
                            link.click()
                            time.sleep(7)
                            soma_files += 1
                            print(f'Baixando {soma_files}/{total_files} arquivos')

                            arquivos_baixados = os.listdir(download_dir)
                            for arquivo in arquivos_baixados:
                                if arquivo.endswith(".zip"):  # Verifica se o arquivo é um ZIP
                                    caminho_arquivo = os.path.join(download_dir, arquivo)
                                    shutil.move(caminho_arquivo, os.path.join(pasta_empresa, arquivo))
                        try:
                            WebDriverWait(driver, 10).until(
                                EC.presence_of_element_located(
                                    (By.XPATH, '//*[@id="DataTables_Table_0_paginate"]/span/a[2]')))
                            driver.find_element('xpath', '//*[@id="DataTables_Table_0_paginate"]/span/a[2]').click()
                            download_links2 = driver.find_element('xpath', '//*[@id="DataTables_Table_0_paginate"]/span/a[2]').click()
                            total_files2 = (len(download_links2))
                            print(f'Total de arquivos da segunda página: {total_files2}')                                        

                            for link in driver.find_elements(By.CLASS_NAME, 'icone-baixar'):
                                link.click()
                                time.sleep(7)
                                soma_files += 1
                                print(f'Baixando {soma_files2}/{total_files2} arquivos')
                                arquivos_baixados = os.listdir(download_dir)
                                for arquivo in arquivos_baixados:
                                    if arquivo.endswith(".zip"):  # Verifica se o arquivo é um ZIP
                                        caminho_arquivo = os.path.join(download_dir, arquivo)
                                        shutil.move(caminho_arquivo, os.path.join(pasta_empresa, arquivo))

                            loop = False

                            if loop == False:

                                linha_celula = linha[4]
                                if hasattr(linha_celula, 'row'):
                                    linha_atual = linha_celula.row
                                    sheet_empresas[f'H{linha_atual}'] = str(int(soma_files)+int(soma_files2)) + "/" + str(int(total_files)+int(total_files2))
                                    workbook.save(valores['caminho'])

                            print(f'Arquivos da empresa CNPJ: {cnpj} baixados com sucesso!')
                            time.sleep(2)
                            driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/a').click()
                            time.sleep(2)
                        except:
                            print(f'Arquivos da empresa CNPJ: {cnpj} baixados com sucesso!')
                            time.sleep(2)
                            driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/a').click()
                            time.sleep(2)
                            continue

            print("Baixa de arquivos finalizada com sucesso")
            time.sleep(3)
            driver.quit()

        # Condição para realizar as buscas quando o certificado é da própria empresa
        if valores['finalidade'] == '1' and valores['proprio'] is True:
            chrome_options = ChromeOptions()
            chromedriver_path = download_undetected_chromedriver(folder_path, undetected=True, arm=False,
                                                                force_update=True)
            driver = uc.Chrome(driver_executable_path=chromedriver_path, options=chrome_options, headless=False)            
            driver.get(url)
            driver.maximize_window()
            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="login-acoes"]/div[2]/p/button'))
            )
            driver.find_element('xpath', '//*[@id="login-acoes"]/div[2]/p/button').click()

            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="login-certificate"]'))
            )
            driver.find_element('xpath', '//*[@id="login-certificate"]').click()

            print("Selecione o certificado para continuar")

            janela.refresh()
            janela.minimize()

            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="sairAplicacao"]'))
            )
            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="menuDownload"]'))
                )
            driver.find_element('xpath', '//*[@id="menuDownload"]').click()
            driver.find_element('xpath', '//*[@id="menuDownload"]').send_keys(Keys.DOWN + Keys.ENTER)
            WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="TipoPedido"]'))
                )
            # Verifica a data de abertura da empresa para pesquisar a partir dali se posterior a 01/01/2018
            cnpj = driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/p[2]/span[1]').text.strip('-').rstrip()
            print(cnpj)

            cnpj_formatado = cnpj.replace('.', '').replace('/', '').replace('-', '')

            # Monta a URL com o CNPJ
            url = (f"https://www.receitaws.com.br/v1/cnpj/{cnpj_formatado}")

            abertura = None

            try:
                # Faz a requisição HTTP
                response = requests.get(url)
                response.raise_for_status()  # Lança uma exceção se a requisição falhar

                # Converte a resposta para JSON
                data = response.json()
                abertura = data.get('abertura')
                print(f"Data de abertura localizada: {abertura}")

            except requests.exceptions.RequestException as e:
                print("Erro na requisição", f"Erro: {e}")

            # Verifica se a data de abertura é válida antes de tentar converter
            start_date = datetime(2018, 1, 1)

            if abertura:
                try:
                    # Converte a data de abertura para um objeto datetime
                    abertura_dt = datetime.strptime(abertura, '%d/%m/%Y')
                    print("Data de abertura original:", abertura_dt)

                    # Substitui o dia pelo dia 01
                    abertura_dt = abertura_dt.replace(day=1)
                    print("Data de inicio utilizada:", abertura_dt)

                    # Defina start_date para abertura_dt se for posterior a 2018-01-01, caso contrário, use 2018-01-01
                    start_date = max(abertura_dt, datetime(2018, 1, 1))
                    
                except ValueError:
                    print("Erro ao converter a data de abertura.")
                    start_date = datetime(2018, 1, 1)
            else:
                start_date = datetime(2018, 1, 1)                

            # Define as datas para busca
            data_corte = datetime.strptime(driver.find_element(By.CLASS_NAME, 'alert-info').text[58:68], '%d/%m/%Y')

            if valores['dataini']:
                start_date = datetime.strptime(valores['dataini'], '%d/%m/%Y')
            else:
                start_date = datetime(2018, 1, 1)

            end_date = start_date + timedelta(days=30 * int(valores['periodo']))

            loop = True
            while loop:
                if end_date > data_corte:
                    end_date = data_corte
                    loop = False

                start_date_string = start_date.strftime("%d/%m/%Y")
                end_date_string = end_date.strftime("%d/%m/%Y")

                print("Inicio: ", start_date_string, "   Fim: ", end_date_string)

                driver.find_element('xpath', '//*[@id="TipoPedido"]/option[2]').click()
                driver.find_element('xpath', '//*[@id="DataInicial"]').send_keys(start_date_string)
                driver.find_element('xpath', '//*[@id="DataFinal"]').click()
                driver.find_element('xpath', '//*[@id="DataFinal"]').clear()
                driver.find_element('xpath', '//*[@id="DataFinal"]').send_keys(end_date_string)
                driver.find_element('xpath', '//*[@id="btnSalvar"]').click()
                pedido = driver.find_element(By.CLASS_NAME, 'fade-alert').text[2:]
                print(pedido)
                if pedido == 'Solicitação enviada com sucesso.':
                    WebDriverWait(driver, 120).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="conteudo-pagina"]/div[1]/a'))
                    )
                    driver.find_element('xpath', '//*[@id="conteudo-pagina"]/div[1]/a').click()

                    start_date = end_date + timedelta(days=1)
                    end_date = start_date + timedelta(days=30 * int(valores['periodo']))

                elif pedido == 'Pedido não foi aceito. Já existe um pedido do mesmo tipo.':
                    driver.find_element(By.XPATH, '//*[@id="btnCancelarAlteracao"]').click()
                    WebDriverWait(driver, 120).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="conteudo-pagina"]/div[1]/a'))
                    )
                    driver.find_element('xpath', '//*[@id="conteudo-pagina"]/div[1]/a').click()

                    start_date = end_date + timedelta(days=1)
                    end_date = start_date + timedelta(days=30 * int(valores['periodo']))
                                
                elif pedido == 'O limite de solicitações foi alcançado. Somente é permitido 72 (doze) solicitações por dia.':
                    time.sleep(2)
                    driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/a').click()
                    break

            print('Buscas Finalizadas')
            time.sleep(3)
            driver.quit()

        # Condição para fazer o download dos arquivos quando o certificado é próprio
        if valores['finalidade'] == '2' and valores['proprio'] is True:
            chrome_options = ChromeOptions()
            prefs = {'download.default_directory': download_dir}
            chrome_options.add_experimental_option('prefs', prefs)
            chromedriver_path = download_undetected_chromedriver(folder_path, undetected=True, arm=False,
                                                                force_update=True)
            driver = uc.Chrome(driver_executable_path=chromedriver_path, options=chrome_options, headless=False)            
            driver.get(url)
            driver.maximize_window()
            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="login-acoes"]/div[2]/p/button'))
            )
            driver.find_element('xpath', '//*[@id="login-acoes"]/div[2]/p/button').click()

            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="login-certificate"]'))
            )
            driver.find_element('xpath', '//*[@id="login-certificate"]').click()

            print("Selecione o certificado para continuar")

            janela.refresh()
            janela.minimize()

            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="sairAplicacao"]'))
            )
            for linha in sheet_empresas.iter_rows(min_row=int(valores['linhaini']), max_row=int(valores['linhafim'])):
                pasta_empresa = criar_pasta(f'{linha[0].value} - {linha[1].value.rstrip()}')
                WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="menuDownload"]'))
                )
                driver.find_element('xpath', '//*[@id="menuDownload"]').click()
                driver.find_element('xpath', '//*[@id="menuDownload"]').send_keys(
                    Keys.DOWN + Keys.DOWN + Keys.ENTER)
                WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located(
                        (By.XPATH, '//*[@id="conteudo-pagina"]/form/section/div/div[4]/input'))
                )
                driver.find_element('xpath', '//*[@id="conteudo-pagina"]/form/section/div/div[4]/input').click()
                time.sleep(10)

                download_links = driver.find_elements('xpath', '//*[@id="DataTables_Table_0"]/tbody/tr/td/a')
                total_files = (len(download_links))
                print(f'Total de arquivos: {total_files}')
                soma_files = 0

                for link in driver.find_elements(By.CLASS_NAME, 'icone-baixar'):
                    link.click()
                    time.sleep(7)
                    soma_files += 1
                    print(f'Baixando {soma_files}/{total_files} arquivos')
                    arquivos_baixados = os.listdir(download_dir)
                    for arquivo in arquivos_baixados:
                        if arquivo.endswith(".zip"):  # Verifica se o arquivo é um ZIP
                            caminho_arquivo = os.path.join(download_dir, arquivo)
                            shutil.move(caminho_arquivo, os.path.join(pasta_empresa, arquivo))

                try:
                    WebDriverWait(driver, 3).until(
                        EC.presence_of_element_located(
                            (By.XPATH, '//*[@id="DataTables_Table_0_paginate"]/span/a[2]')))
                    driver.find_element('xpath', '//*[@id="DataTables_Table_0_paginate"]/span/a[2]').click()
                    download_links = driver.find_elements('xpath', '//*[@id="DataTables_Table_0"]/tbody/tr/td/a')
                    total_files2 = (len(download_links))
                    for link in driver.find_elements(By.CLASS_NAME, 'icone-baixar'):
                        link.click()
                        time.sleep(7)
                        soma_files += 1
                        print(f'Baixando {soma_files}/{total_files + total_files2} arquivos')
                        arquivos_baixados = os.listdir(download_dir)
                        for arquivo in arquivos_baixados:
                            if arquivo.endswith(".zip"):  # Verifica se o arquivo é um ZIP
                                caminho_arquivo = os.path.join(download_dir, arquivo)
                                shutil.move(caminho_arquivo, os.path.join(pasta_empresa, arquivo))

                except:
                    time.sleep(2)
                    driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/a').click()
                    time.sleep(2)
                    continue

                time.sleep(2)
                driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/a').click()
                time.sleep(2)

            print("Baixa de arquivos finalizada com sucesso")
            time.sleep(5)

        print("Programa Finalizado")
        time.sleep(5)
        driver.quit()                    