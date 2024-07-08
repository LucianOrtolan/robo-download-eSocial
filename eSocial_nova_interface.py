import time
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import undetected_chromedriver as uc
from auto_download_undetected_chromedriver import download_undetected_chromedriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from undetected_chromedriver import ChromeOptions
import openpyxl
from openpyxl.utils.cell import get_column_letter
import os
import shutil
import requests
from bs4 import BeautifulSoup
from threading import Thread
from tkinter import messagebox
import psutil
import signal
from scriptdownload import script_download

def retorna_qtd_loop(start_date,end_date,data_corte):
      loop = True
      data_ini = start_date
      data_fim = end_date
      conta_loop = 0 
       
      while loop:             
         if data_fim > data_corte:
            data_fim = data_corte
            loop = False   
             
         data_ini = data_fim + timedelta(days=1)
         data_fim = data_ini + timedelta(days=31*int(meses_buscar_var.get()))
         conta_loop = conta_loop + 1 
          
      return conta_loop

def ret_data_abertura_empresa(cnpj):      
      cnpj_sem_mascara = ''.join(filter(str.isdigit, cnpj))
      # Faz a solicitação HTTP para obter o JSON
      try:
         response = requests.get(f"https://receitaws.com.br/v1/cnpj/{cnpj_sem_mascara}")         
         abertura_data = datetime(2018, 1, 1) 
         # Verifica se a solicitação foi bem-sucedida (código de status 200)
         if response.status_code == 200:
            # Converte a resposta JSON em um dicionário Python
            data_json = response.json()       
            # Obtém o valor da chave "abertura"
            abertura_string = data_json.get("abertura")           
            abertura_data = datetime.strptime(abertura_string, "%d/%m/%Y")
            print(f"Data de abertura: {abertura_data.strftime('%d/%m/%Y')}")
            #print("abertura_data CNPJ:",abertura_data) 
         else:
            print('Sem retorno de data na rotina 1') 
            response = requests.get(f"https://api-publica.speedio.com.br/buscarcnpj?cnpj={cnpj_sem_mascara}")             
            abertura_data = datetime(2018, 1, 1)                       
            if response.status_code == 200:
               data_json = response.json()  
               abertura_string = data_json.get("DATA ABERTURA")              
               abertura_data = datetime.strptime(abertura_string, "%d/%m/%Y")
               print(f"Data de abertura: {abertura_data.strftime('%d/%m/%Y')}")
            else:              
               print("Sem retorno de data na rotina 2") 
               response = requests.get(f"https://brasilapi.com.br/api/cnpj/v1/{cnpj_sem_mascara}") 
                
               abertura_data = datetime(2018, 1, 1)               
               if response.status_code == 200:
                  data_json = response.json()    
                  abertura_string = data_json.get("data_inicio_atividade")                  
                  abertura_data = datetime.strptime(abertura_string, "%Y-%m-%d")
                  print(f"Data de abertura: {abertura_data.strftime('%d/%m/%Y')}") 
               else:              
                  print("Sem retorno de data na rotina 3") 
       
      except:
         abertura_data = datetime(2018, 1, 1)
         print("Erro na consulta do CNPJ") 

      response.close()
       
      data_inicio_padrao = datetime(2018, 1, 1)
      if data_inicio_padrao > abertura_data:
         abertura_data = data_inicio_padrao

      d_abertura_string = abertura_data.strftime("%d/%m/%Y")       
      return abertura_data 
class BarraProgresso:
    def __init__(self):
        self.barra_window = tk.Toplevel()
        self.barra_window.title("Progresso do envio das datas")
        self.barra_window.geometry("350x150")
        self.barra_window.resizable(False, False)
        self.barra_window.lift()  # Colocando a barra de progresso acima da janela principal
        self.barra_window.focus_force()
        
        # Calcula a posição para centralizar a janela
        largura_janela = self.barra_window.winfo_reqwidth()
        altura_janela = self.barra_window.winfo_reqheight()
        largura_tela = self.barra_window.winfo_screenwidth()
        altura_tela = self.barra_window.winfo_screenheight()
        x_pos = (largura_tela - largura_janela) // 2
        y_pos = (altura_tela - altura_janela) // 2
        self.barra_window.geometry(f"+{x_pos}+{y_pos}")
        
        self.progressbar = ttk.Progressbar(self.barra_window, orient="horizontal", length=200, mode="determinate")
        self.progressbar.pack(side="top", padx=10, pady=10)
        self.porcentagem_label = tk.Label(self.barra_window, text="", font=("Arial", 12))
        self.porcentagem_label.pack(side="top", padx=10, pady=5)
        self.descricao_datas = tk.Label(self.barra_window, text="", font=("Arial", 12))
        self.descricao_datas.pack(side="top", padx=10, pady=5)
        self.descricao_label = tk.Label(self.barra_window, text="", font=("Arial", 12))
        self.descricao_label.pack(side="top", padx=10, pady=5)

    def mostrar_barra(self, conta_loop,qtd_loop,start_date_string,end_date_string):
        porcentagem = conta_loop * 100 // qtd_loop
        self.progressbar["value"] = porcentagem
        self.porcentagem_label.config(text=f"{porcentagem}%")
        self.descricao_datas.config(text=f"{start_date_string} até {end_date_string}")
        self.descricao_label.config(text=f"{conta_loop} de {qtd_loop}")
        self.barra_window.update_idletasks()  # Atualiza a tela
        self.progressbar.update()

    def fechar_barra(self):
        self.barra_window.destroy()

def find_and_kill_process(process_name):
    # Itera sobre todos os processos em execução
    for proc in psutil.process_iter(attrs=['pid', 'name']):
        if process_name.lower() in proc.info['name'].lower():
            # Se o processo for encontrado, mata o processo
            os.kill(proc.info['pid'], signal.SIGTERM)
            print(f'Processo {process_name} (PID: {proc.info["pid"]}) foi encerrado.')
            return
    print(f'Processo {process_name} não encontrado.')

find_and_kill_process('chromedriver.exe')

def solicitar_ou_baixar():
    # Função que será chamada quando o botão for clicado
    opcao = solicitar_baixar_var.get()
    find_and_kill_process('chromedriver.exe')    
    if opcao == 1 and certificado_proprio_var.get() is False:        
        url = 'https://login.esocial.gov.br/login.aspx'
        folder_path = "c:\\chromedriver"
        chrome_options = ChromeOptions()
        chromedriver_path = download_undetected_chromedriver(folder_path, undetected=True, arm=False,
                                                                        force_update=True)
        driver = uc.Chrome(driver_executable_path=chromedriver_path, options=chrome_options, headless=False)            
        driver.get(url)
        driver.maximize_window()
        root.iconify()
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="login-acoes"]/div[2]/p/button'))
        )
        driver.find_element('xpath', '//*[@id="login-acoes"]/div[2]/p/button').click()

        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="login-certificate"]'))
        )
        driver.find_element('xpath', '//*[@id="login-certificate"]').click()

        print("Selecione o certificado para continuar")        

        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="sairAplicacao"]'))
        )
        
        inscricao = driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/p[2]/span[1]').text.strip('-')
        print(f'CNPJ do procurador: {inscricao}')
        # Condição para identificar se a inscrição é um CNPJ ou CPF
        if len(driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/p[2]/span[1]').text) < 18:
            driver.find_element(By.XPATH, '//*[@id="geral"]/div').click()
            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="header"]/div[2]/a'))
            )
            driver.find_element('xpath', '//*[@id="header"]/div[2]/a').click()
            pass
        else:
            driver.find_element(By.CLASS_NAME, 'alterar-perfil').click()
        
        workbook = openpyxl.load_workbook(caminho_planilha_var.get())
        sheet_empresas = workbook.active        
        documento = sheet_empresas.cell(row=int(linha_ini.get()), column=3).value      
        
        data_corte = ''                   
        loop = True
        erros_datas = ''

        for linha in sheet_empresas.iter_rows(min_row=int(linha_ini.get()), max_row=int(linha_fim.get())):
            documento = len(str(linha[2].value))            
            # Condição que verifica se é CNPJ ou CPF na planilha                                
            if documento >= 15:
                cnpj = linha[2].value # CNPJ                
                driver.find_element('xpath', '//*[@id="perfilAcesso"]').click()
                driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.DOWN + Keys.DOWN)
                driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.ENTER)
                driver.find_element('xpath', '//*[@id="procuradorCnpj"]').send_keys(cnpj)
                driver.find_element('xpath', '//*[@id="procuradorCnpj"]').send_keys(Keys.LEFT_CONTROL + 'v')
                WebDriverWait(driver, 120).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="btn-verificar-procuracao-cnpj"]'))
                    )
                driver.find_element('xpath', '//*[@id="btn-verificar-procuracao-cnpj"]').click()
                mensagem_procuracao = ''
                cnpjInvalido = driver.find_element(By.XPATH, '//*[@id="procuradorCnpj-error"]').text[0:]
                try:
                    WebDriverWait(driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                    )
                except:
                    mensagem_procuracao = driver.find_element(By.CLASS_NAME, 'fade-alert').text[2:]
                    cnpjInvalido

                # Condição se verifica se possui procuração para o CNPJ que está sendo procurado
                print(mensagem_procuracao)
                if mensagem_procuracao == 'O procurador não possui perfil com autorização de acesso à Web':
                    print(f'Não possui procuração para o {cnpj}')
                    linha_celula = linha[4]

                    if hasattr(linha_celula, 'row'):
                        linha_atual = linha_celula.row
                        sheet_empresas[f"H{linha_atual}"] = "Não possui procuração"                        
                        workbook.save(caminho_planilha_var.get())                        
                        print('Retornando as buscas')
                        driver.refresh()
                        continue

                elif cnpjInvalido == 'CNPJ inválido.':                    
                    print(f'CNPJ inválido {cnpj}')
                    linha_celula = linha[4]
                    
                    if hasattr(linha_celula, 'row'):
                        linha_atual = linha_celula.row
                        sheet_empresas[f"H{linha_atual}"] = "CNPJ inválido"                        
                        workbook.save(caminho_planilha_var.get())                        
                        print('Retornando as buscas')
                        driver.refresh()
                        continue
                
                else:
                    print(f'CNPJ/CPF sendo buscado: {str(linha[2].value)}')
                    driver.find_element('xpath', '//*[@id="geral"]/div').click()
                    WebDriverWait(driver, 120).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="menuDownload"]'))
                    )
                    driver.find_element('xpath', '//*[@id="menuDownload"]').click()
                    driver.find_element('xpath', '//*[@id="menuDownload"]').send_keys(Keys.DOWN + Keys.ENTER)
                    WebDriverWait(driver, 120).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="TipoPedido"]'))
                    )
                    
                    data_abertura = ret_data_abertura_empresa(cnpj)
                    data_inicial = datetime(2018, 1, 1)
                    if data_abertura >= data_inicial:                        
                        start_date = data_abertura
                    else:
                        start_date = data_inicial

                    data_final = start_date + timedelta(days=31*int(meses_buscar_var.get()))    
                    data_corte = datetime.strptime(driver.find_element(By.CLASS_NAME, 'alert-info').text[58:68],'%d/%m/%Y')

                    qtd_loop = retorna_qtd_loop(start_date,data_final,data_corte)
                    conta_loop = 0 
                    erro_script = False
                    loop = True

                    barra_progresso = BarraProgresso()
                    barra_thread = Thread()              
                    barra_thread.start() 

                    # Faz as requisições para a empresa atual
                    while loop:
                        if data_final > data_corte:
                            data_final = data_corte
                            loop = False

                        data_inicial_str = start_date.strftime('%d/%m/%Y')
                        data_final_str = data_final.strftime('%d/%m/%Y')

                        print(f'Data Inicial: {data_inicial_str} - Data Final: {data_final_str}')

                        conta_loop = conta_loop + 1
                        barra_progresso.mostrar_barra(conta_loop,qtd_loop,data_inicial_str,data_final_str)

                        url = 'https://www.esocial.gov.br/portal/download/Pedido/Solicitacao' 
                        dados = {
                                "npjOperadorPortuario": "",
                                "CodigoLotacao": "",
                                "CodigoRubrica": "",
                                "CpfTrabalhador": "",
                                "DataFinal": data_final_str,
                                "DataInicial": data_inicial_str,
                                "HoraFinal": "23",
                                "HoraInicial": "00",
                                "IdTabelaRubrica": "",
                                "NumeroProcesso": "",
                                "PerApur": "",
                                "TipoPedido": "1",
                                "TipoProcesso": "0"
                                }
                        
                        response = driver.execute_script(f"""
                            async function fetchData() {{
                                try {{
                                    const response = await fetch('{url}', {{ method: 'POST', headers: {{ 'Content-Type': 'application/json' }}, body: JSON.stringify({dados}), redirect: 'manual' }});
                                    const contentType = response.headers.get('content-type');
                                    const data = await response.text();
                                    return {{ data: data, contentType: contentType }};
                                }} catch (error) {{
                                    return {{ error: error.message }};
                                }}
                            }}
                            return fetchData();
                        """)
                        
                        # Processa a resposta
                        if 'error' in response:
                            barra_progresso.fechar_barra()
                            print("Erro durante a solicitação:", response['error'])                                
                            break          
                        else:
                            content_type = response['contentType']
                            if content_type and 'text/html' in content_type:
                                html_content = response['data']
                                if html_content.strip():
                                    # Analisa o conteúdo HTML
                                    soup = BeautifulSoup(html_content, 'html.parser')
                                    # Procura por divs com a classe específica
                                    alert_divs = soup.find_all('div', class_='fade-alert alert alert-danger retornoServidor')
                                    # Verifica se foram encontradas divs
                                    if alert_divs:
                                        for div in alert_divs:
                                            print("Erro encontrado:", div.text.strip())
                                            erro_script = True
                                            erro_script_todas_empresas = True
                                            erros_datas = erros_datas + data_inicial_str + ' a ' + data_final_str + ' - ' + div.text.strip() + '\n\n' 

                                    else:
                                        print("Nenhuma mensagem encontrada na página.")
                                else:
                                    print("A resposta HTML está vazia.")
                    
                        # Atualiza a data atual para a próxima iteração                            
                        start_date = data_final + timedelta(days=1)
                        data_final = start_date + timedelta(days=31*int(meses_buscar_var.get()))
                        linha_celula = linha[4]
                        if hasattr(linha_celula, 'row'):
                            linha_atual = linha_celula.row
                            sheet_empresas[f"H{linha_atual}"] = "OK"
                            sheet_empresas[f"J{linha_atual}"] = data_inicial_str + " a " + data_final_str
                            workbook.save(caminho_planilha_var.get())
                    
                    barra_progresso.fechar_barra()
                    driver.find_element('xpath', '//*[@id="header"]/div[2]/a').click()
            else:
                # Buscas por CPF
                cnpj = linha[2].value                
                driver.find_element('xpath', '//*[@id="perfilAcesso"]').click()
                driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.DOWN)
                driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.ENTER)
                driver.find_element('xpath', '//*[@id="procuradorCpf"]').send_keys(cnpj)
                driver.find_element('xpath', '//*[@id="procuradorCpf"]').send_keys(Keys.LEFT_CONTROL + 'v')
                WebDriverWait(driver, 120).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="btn-verificar-procuracao-cpf"]'))
                    )
                driver.find_element('xpath', '//*[@id="btn-verificar-procuracao-cpf"]').click()
                mensagem_procuracao = ''
                try:
                    WebDriverWait(driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                    )
                except:
                    mensagem_procuracao = driver.find_element(By.CLASS_NAME, 'fade-alert').text[2:]
                
                # Condição se verifica se possui procuração para o CNPJ que está sendo procurado
                print(mensagem_procuracao)
                if mensagem_procuracao == 'O procurador não possui perfil com autorização de acesso à Web':
                    print(f'Não possui procuração para o {cnpj}')
                    linha_celula = linha[4]

                    if hasattr(linha_celula, 'row'):
                        linha_atual = linha_celula.row
                        sheet_empresas[f'F{linha_atual}'] = 'Não possui procuração'
                        workbook.save(caminho_planilha_var.get())
                        print('Retornando as buscas')
                        driver.refresh()
                        continue
                else:
                    print(f'CNPJ/CPF sendo buscado: {str(linha[2].value)}')
                    driver.find_element('xpath', '//*[@id="geral"]/div').click()
                    WebDriverWait(driver, 120).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="menuDownload"]'))
                    )
                    driver.find_element('xpath', '//*[@id="menuDownload"]').click()
                    driver.find_element('xpath', '//*[@id="menuDownload"]').send_keys(Keys.DOWN + Keys.ENTER)
                    WebDriverWait(driver, 120).until(
                        EC.presence_of_element_located((By.XPATH, '//*[@id="TipoPedido"]'))
                    )
                    
                    data_abertura = ret_data_abertura_empresa(cnpj)
                    data_inicial = datetime(2018, 1, 1)
                    if data_abertura >= data_inicial:                        
                        start_date = data_abertura
                    else:
                        start_date = data_inicial

                    data_final = start_date + timedelta(days=31*int(meses_buscar_var.get()))    
                    data_corte = datetime.strptime(driver.find_element(By.CLASS_NAME, 'alert-info').text[58:68],'%d/%m/%Y')

                    qtd_loop = retorna_qtd_loop(start_date,data_final,data_corte)
                    conta_loop = 0 
                    erro_script = False
                    loop = True

                    barra_progresso = BarraProgresso()
                    barra_thread = Thread()              
                    barra_thread.start() 

                    # Faz as requisições para a empresa atual
                    while loop:
                        if data_final > data_corte:
                            data_final = data_corte
                            loop = False

                        data_inicial_str = start_date.strftime('%d/%m/%Y')
                        data_final_str = data_final.strftime('%d/%m/%Y')

                        print(f'Data Inicial: {data_inicial_str} - Data Final: {data_final_str}')

                        conta_loop = conta_loop + 1
                        barra_progresso.mostrar_barra(conta_loop,qtd_loop,data_inicial_str,data_final_str)

                        url = 'https://www.esocial.gov.br/portal/download/Pedido/Solicitacao' 
                        dados = {
                                "npjOperadorPortuario": "",
                                "CodigoLotacao": "",
                                "CodigoRubrica": "",
                                "CpfTrabalhador": "",
                                "DataFinal": data_final_str,
                                "DataInicial": data_inicial_str,
                                "HoraFinal": "23",
                                "HoraInicial": "00",
                                "IdTabelaRubrica": "",
                                "NumeroProcesso": "",
                                "PerApur": "",
                                "TipoPedido": "1",
                                "TipoProcesso": "0"
                                }
                        
                        response = driver.execute_script(f"""
                            async function fetchData() {{
                                try {{
                                    const response = await fetch('{url}', {{ method: 'POST', headers: {{ 'Content-Type': 'application/json' }}, body: JSON.stringify({dados}), redirect: 'manual' }});
                                    const contentType = response.headers.get('content-type');
                                    const data = await response.text();
                                    return {{ data: data, contentType: contentType }};
                                }} catch (error) {{
                                    return {{ error: error.message }};
                                }}
                            }}
                            return fetchData();
                        """)
                        erros_datas = ''
                        # Processa a resposta
                        if 'error' in response:
                            barra_progresso.fechar_barra()
                            print("Erro durante a solicitação:", response['error'])                                
                            break          
                        else:
                            content_type = response['contentType']
                            if content_type and 'text/html' in content_type:
                                html_content = response['data']
                                if html_content.strip():
                                    # Analisa o conteúdo HTML
                                    soup = BeautifulSoup(html_content, 'html.parser')
                                    # Procura por divs com a classe específica
                                    alert_divs = soup.find_all('div', class_='fade-alert alert alert-danger retornoServidor')
                                    # Verifica se foram encontradas divs
                                    if alert_divs:
                                        for div in alert_divs:
                                            print("Erro encontrado:", div.text.strip())
                                            erro_script = True
                                            erro_script_todas_empresas = True
                                            erros_datas = erros_datas + data_inicial_str + ' a ' + data_final_str + ' - ' + div.text.strip() + '\n\n' 

                                    else:
                                        print("Nenhuma mensagem encontrada na página.")
                                else:
                                    print("A resposta HTML está vazia.")
                    
                        # Atualiza a data atual para a próxima iteração                            
                        start_date = data_final + timedelta(days=1)
                        data_final = start_date + timedelta(days=31*int(meses_buscar_var.get()))
                        linha_celula = linha[4]
                        if hasattr(linha_celula, 'row'):
                            linha_atual = linha_celula.row
                            sheet_empresas[f"H{linha_atual}"] = "OK"
                            sheet_empresas[f"J{linha_atual}"] = data_inicial_str + " a " + data_final_str
                            workbook.save(caminho_planilha_var.get())
                    
                    barra_progresso.fechar_barra()
                    driver.find_element('xpath', '//*[@id="header"]/div[2]/a').click()
        
        print('Buscas Finalizadas')
        time.sleep(7)
        driver.quit()        

    elif opcao == 2 and certificado_proprio_var.get() is False:
        url = 'https://login.esocial.gov.br/login.aspx'
        folder_path = "c:\\chromedriver"
        download_dir = caminho_pasta_salvar_var.get().rstrip().replace('/', '\\')

        def criar_pasta(nome_empresa):
            pasta_empresa = os.path.join(download_dir, nome_empresa)
            if not os.path.exists(pasta_empresa):
                os.makedirs(pasta_empresa)
            return pasta_empresa
        
        chrome_options = ChromeOptions()
        prefs = {"download.default_directory": download_dir,
                 "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True,
                 "profile.default_content_setting_values.automatic_downloads": 1,
                 "profile.content_settings.exceptions.automatic_downloads.*.setting": 1,
                 "safebrowsing.enabled": False}
        chrome_options.add_experimental_option('prefs', prefs)
        chromedriver_path = download_undetected_chromedriver(folder_path, undetected=True, arm=False,
                                                            force_update=True)
        driver = uc.Chrome(driver_executable_path=chromedriver_path, options=chrome_options,headless=False)            
        driver.get(url)
        driver.maximize_window()
        root.iconify()
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="login-acoes"]/div[2]/p/button'))
        )
        driver.find_element('xpath', '//*[@id="login-acoes"]/div[2]/p/button').click()

        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="login-certificate"]'))
        )
        driver.find_element('xpath', '//*[@id="login-certificate"]').click()

        print("Selecione o certificado para continuar")        

        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="sairAplicacao"]'))
        )            
        inscricao = driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/p[2]/span[1]').text.strip('-')
        print(f'CNPJ do procurador: {inscricao}')
        if len(driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/p[2]/span[1]').text) < 18:
            driver.find_element(By.XPATH, '//*[@id="geral"]/div').click()
            WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="header"]/div[2]/a'))
            )
            driver.find_element('xpath', '//*[@id="header"]/div[2]/a').click()
            pass
        else:
            driver.find_element(By.CLASS_NAME, 'alterar-perfil').click()
        
        workbook = openpyxl.load_workbook(caminho_planilha_var.get())
        sheet_empresas = workbook.active        
        documento = sheet_empresas.cell(row=int(linha_ini.get()), column=3).value        
        empresas = list(sheet_empresas.iter_rows(min_row=int(linha_ini.get()), max_row=int(linha_fim.get())))            

        for linha in sheet_empresas.iter_rows(min_row=int(linha_ini.get()), max_row=int(linha_fim.get())):
            WebDriverWait(driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="perfilAcesso"]'))
                    )
            pasta_empresa = criar_pasta(f'{linha[0].value} - {linha[1].value.rstrip()}')
            documento = len(str(linha[2].value))
            if documento >= 15: #CNPJ
                cnpj = linha[2].value
                driver.find_element('xpath', '//*[@id="perfilAcesso"]').click()
                driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.DOWN + Keys.DOWN)
                driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.ENTER)
                driver.find_element('xpath', '//*[@id="procuradorCnpj"]').send_keys(cnpj)
                driver.find_element('xpath', '//*[@id="procuradorCnpj"]').send_keys(Keys.LEFT_CONTROL + 'v')
                WebDriverWait(driver, 120).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="btn-verificar-procuracao-cnpj"]'))
                        )
                driver.find_element('xpath', '//*[@id="btn-verificar-procuracao-cnpj"]').click()
                mensagem_procuracao = ''
                try:
                    WebDriverWait(driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                    )
                except:
                    mensagem_procuracao = driver.find_element(By.CLASS_NAME, 'fade-alert').text[2:]

                # Condição se verifica se possui procuração para o CNPJ que está sendo procurado                                
                if mensagem_procuracao:
                    print(f'Não possui procuração para o {cnpj}')
                    linha_celula = linha[4]

                    if hasattr(linha_celula, 'row'):
                        linha_atual = linha_celula.row
                        sheet_empresas[f'H{linha_atual}'] = 'Não possui procuração'
                        workbook.save(caminho_planilha_var.get())
                        print('Retornando as buscas')
                        driver.refresh()
                        continue
                else:
                    WebDriverWait(driver, 120).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                    )
                    driver.find_element('xpath', '//*[@id="geral"]/div').click()
                    WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="menuDownload"]'))
                )
                    driver.get("https://www.esocial.gov.br/portal/download/Pedido/Consulta")
                    WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="conteudo-pagina"]/form/section/div/div[4]/input')))
                    driver.find_element(By.XPATH, '//*[@id="conteudo-pagina"]/form/section/div/div[4]/input').click()
                    driver.execute_script(script_download)
                    print("Baixando arquivos...")                        
                    WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[13]/button[2]')))
                    driver.find_element(By.XPATH, '/html/body/div[13]/button[2]').click()
                    # Função para mover arquivos .zip
                    def verificar_e_mover_arquivos_zip():
                        arquivos_baixados = os.listdir(download_dir)
                        for arquivo in arquivos_baixados:
                            if arquivo.endswith(".zip"):  # Verifica se o arquivo é um ZIP
                                caminho_arquivo = os.path.join(download_dir, arquivo)
                                shutil.move(caminho_arquivo, os.path.join(pasta_empresa, arquivo))
                                print(f"Arquivo {arquivo} movido para {pasta_empresa}")
                    
                    intervalo_verificacao = 20  # segundos
                    time.sleep(10)
                    download_dir = caminho_pasta_salvar_var.get().rstrip().replace('/', '\\')            
                    loop = True            
                    while loop == True:
                        arquivos_baixados = os.listdir(download_dir)
                        existe_zip = any(arquivo.endswith('.zip') for arquivo in arquivos_baixados)
                        if existe_zip:
                            verificar_e_mover_arquivos_zip()
                            print(f"Aguardando {intervalo_verificacao} segundos para próxima verificação de arquivos baixados...")
                            time.sleep(intervalo_verificacao)                    
                        else:
                            loop = False
                            print(f"Todos arquivos da empresa {cnpj} foram baixados!")

            else: #CPF
                cnpj = linha[2].value
                driver.find_element('xpath', '//*[@id="perfilAcesso"]').click()
                driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.DOWN)
                driver.find_element('xpath', '//*[@id="perfilAcesso"]').send_keys(Keys.ENTER)
                driver.find_element('xpath', '//*[@id="procuradorCpf"]').send_keys(cnpj)
                driver.find_element('xpath', '//*[@id="procuradorCpf"]').send_keys(Keys.LEFT_CONTROL + 'v')
                WebDriverWait(driver, 120).until(
                            EC.element_to_be_clickable((By.XPATH, '//*[@id="btn-verificar-procuracao-cpf"]'))
                        )
                driver.find_element('xpath', '//*[@id="btn-verificar-procuracao-cpf"]').click()
                mensagem_procuracao = ''
                try:
                    WebDriverWait(driver, 15).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                    )
                except:
                    mensagem_procuracao = driver.find_element(By.CLASS_NAME, 'fade-alert').text[2:]

                # Condição se verifica se possui procuração para o CNPJ que está sendo procurado                                
                if mensagem_procuracao:
                    print(f'Não possui procuração para o {cnpj}')
                    linha_celula = linha[4]

                    if hasattr(linha_celula, 'row'):
                        linha_atual = linha_celula.row
                        sheet_empresas[f'H{linha_atual}'] = 'Não possui procuração'
                        workbook.save(caminho_planilha_var.get())
                        print('Retornando as buscas')
                        driver.refresh()
                        continue
                else:
                    WebDriverWait(driver, 120).until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                    )
                    driver.find_element('xpath', '//*[@id="geral"]/div').click()
                    WebDriverWait(driver, 120).until(
                    EC.presence_of_element_located((By.XPATH, '//*[@id="menuDownload"]'))
                )
                    driver.get("https://www.esocial.gov.br/portal/download/Pedido/Consulta")
                    WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="conteudo-pagina"]/form/section/div/div[4]/input')))
                    driver.find_element(By.XPATH, '//*[@id="conteudo-pagina"]/form/section/div/div[4]/input').click()
                    driver.execute_script(script_download)
                    print("Baixando arquivos...")                        
                    WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[13]/button[2]')))
                    driver.find_element(By.XPATH, '/html/body/div[13]/button[2]').click()
                    # Função para mover arquivos .zip
                    def verificar_e_mover_arquivos_zip():
                        arquivos_baixados = os.listdir(download_dir)
                        for arquivo in arquivos_baixados:
                            if arquivo.endswith(".zip"):  # Verifica se o arquivo é um ZIP
                                caminho_arquivo = os.path.join(download_dir, arquivo)
                                shutil.move(caminho_arquivo, os.path.join(pasta_empresa, arquivo))
                                print(f"Arquivo {arquivo} movido para {pasta_empresa}")
                    
                    intervalo_verificacao = 20  # segundos
                    time.sleep(10)
                    download_dir = caminho_pasta_salvar_var.get().rstrip().replace('/', '\\')            
                    loop = True            
                    while loop == True:
                        arquivos_baixados = os.listdir(download_dir)
                        existe_zip = any(arquivo.endswith('.zip') for arquivo in arquivos_baixados)
                        if existe_zip:
                            verificar_e_mover_arquivos_zip()
                            print(f"Aguardando {intervalo_verificacao} segundos para próxima verificação de arquivos baixados...")
                            time.sleep(intervalo_verificacao)                    
                        else:
                            loop = False
                            print(f"Todos arquivos da empresa {cnpj} foram baixados!")

            driver.find_element(By.XPATH, '//*[@id="header"]/div[2]/a').click()
            WebDriverWait(driver, 120).until(
                    EC.element_to_be_clickable((By.XPATH, '//*[@id="geral"]/div'))
                )

        print("Baixa de arquivos finalizada!")
        time.sleep(3)
        driver.quit()    

    elif opcao == 1 and certificado_proprio_var.get():
        url = 'https://login.esocial.gov.br/login.aspx'
        folder_path = "c:\\chromedriver"
        chrome_options = ChromeOptions()
        chromedriver_path = download_undetected_chromedriver(folder_path, undetected=True, arm=False,
                                                            force_update=True)
        driver = uc.Chrome(driver_executable_path=chromedriver_path, options=chrome_options, headless=False)            
        driver.get(url)
        driver.maximize_window()
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="login-acoes"]/div[2]/p/button'))
        )
        driver.find_element('xpath', '//*[@id="login-acoes"]/div[2]/p/button').click()

        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="login-certificate"]'))
        )
        driver.find_element('xpath', '//*[@id="login-certificate"]').click()

        print("Selecione o certificado para continuar")

        root.iconify()

        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="sairAplicacao"]'))
        )
        WebDriverWait(driver, 120).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="menuDownload"]'))
            )
        driver.find_element('xpath', '//*[@id="menuDownload"]').click()
        driver.find_element('xpath', '//*[@id="menuDownload"]').send_keys(Keys.DOWN + Keys.ENTER)
        WebDriverWait(driver, 120).until(
                EC.presence_of_element_located((By.XPATH, '//*[@id="TipoPedido"]'))
            )
        cnpj = driver.find_element('xpath', '//*[@id="header"]/div[2]/p[2]/span[1]').text[0:18]        
        data_abertura = ret_data_abertura_empresa(cnpj)
        data_inicial = datetime(2018, 1, 1)
        
        if data_abertura >= data_inicial:
            start_date = data_abertura
        else:
            start_date = data_inicial

        data_final = start_date + timedelta(days=31*int(meses_buscar_var.get()))    
        data_corte = datetime.strptime(driver.find_element(By.CLASS_NAME, 'alert-info').text[58:68],'%d/%m/%Y')

        qtd_loop = retorna_qtd_loop(start_date,data_final,data_corte)
        conta_loop = 0 
        erro_script = False
        loop = True

        barra_progresso = BarraProgresso()
        barra_thread = Thread()              
        barra_thread.start() 

        # Faz as requisições para a empresa atual
        while loop:
            if data_final > data_corte:
                data_final = data_corte
                loop = False

            data_inicial_str = start_date.strftime('%d/%m/%Y')
            data_final_str = data_final.strftime('%d/%m/%Y')

            print(f'Data Inicial: {data_inicial_str} - Data Final: {data_final_str}')

            conta_loop = conta_loop + 1
            barra_progresso.mostrar_barra(conta_loop,qtd_loop,data_inicial_str,data_final_str)

            url = 'https://www.esocial.gov.br/portal/download/Pedido/Solicitacao' 
            dados = {
                    "npjOperadorPortuario": "",
                    "CodigoLotacao": "",
                    "CodigoRubrica": "",
                    "CpfTrabalhador": "",
                    "DataFinal": data_final_str,
                    "DataInicial": data_inicial_str,
                    "HoraFinal": "23",
                    "HoraInicial": "00",
                    "IdTabelaRubrica": "",
                    "NumeroProcesso": "",
                    "PerApur": "",
                    "TipoPedido": "1",
                    "TipoProcesso": "0"
                    }
            
            response = driver.execute_script(f"""
                async function fetchData() {{
                    try {{
                        const response = await fetch('{url}', {{ method: 'POST', headers: {{ 'Content-Type': 'application/json' }}, body: JSON.stringify({dados}), redirect: 'manual' }});
                        const contentType = response.headers.get('content-type');
                        const data = await response.text();
                        return {{ data: data, contentType: contentType }};
                    }} catch (error) {{
                        return {{ error: error.message }};
                    }}
                }}
                return fetchData();
            """)
            
            erros_datas = ''
            # Processa a resposta
            if 'error' in response:
                barra_progresso.fechar_barra()
                print("Erro durante a solicitação:", response['error'])                                
                break          
            else:
                content_type = response['contentType']
                if content_type and 'text/html' in content_type:
                    html_content = response['data']
                    if html_content.strip():
                        # Analisa o conteúdo HTML
                        soup = BeautifulSoup(html_content, 'html.parser')
                        # Procura por divs com a classe específica
                        alert_divs = soup.find_all('div', class_='fade-alert alert alert-danger retornoServidor')
                        # Verifica se foram encontradas divs
                        if alert_divs:
                            for div in alert_divs:
                                print("Erro encontrado:", div.text.strip())
                                erro_script = True
                                erro_script_todas_empresas = True
                                erros_datas = erros_datas + data_inicial_str + ' a ' + data_final_str + ' - ' + div.text.strip() + '\n\n' 

                        else:
                            print("Nenhuma mensagem encontrada na página.")
                    else:
                        print("A resposta HTML está vazia.")
        
            # Atualiza a data atual para a próxima iteração                            
            start_date = data_final + timedelta(days=1)
            data_final = start_date + timedelta(days=31*int(meses_buscar_var.get()))
                    
        barra_progresso.fechar_barra()        

        print('Buscas Finalizadas')
        time.sleep(3)
        driver.quit()

    elif opcao == 2 and certificado_proprio_var.get():
        url = 'https://login.esocial.gov.br/login.aspx'
        folder_path = "c:\\chromedriver"
        download_dir = caminho_pasta_salvar_var.get().rstrip().replace('/', '\\')
        
        def criar_pasta(nome_empresa):
            pasta_empresa = os.path.join(download_dir, nome_empresa)
            if not os.path.exists(pasta_empresa):
                os.makedirs(pasta_empresa)
            return pasta_empresa

        chrome_options = ChromeOptions()
        prefs = {"download.default_directory": download_dir,
                 "download.prompt_for_download": False,
                "download.directory_upgrade": True,
                "plugins.always_open_pdf_externally": True,
                 "profile.default_content_setting_values.automatic_downloads": 1,
                 "profile.content_settings.exceptions.automatic_downloads.*.setting": 1,
                 "safebrowsing.enabled": False}
        chrome_options.add_experimental_option('prefs', prefs)
        chromedriver_path = download_undetected_chromedriver(folder_path, undetected=True, arm=False,
                                                            force_update=True)
        driver = uc.Chrome(driver_executable_path=chromedriver_path, options=chrome_options, headless=False)            
        driver.get(url)
        driver.maximize_window()
        WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, '//*[@id="login-acoes"]/div[2]/p/button')))
        driver.find_element('xpath', '//*[@id="login-acoes"]/div[2]/p/button').click()
        WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, '//*[@id="login-certificate"]')))
        driver.find_element('xpath', '//*[@id="login-certificate"]').click()
        print("Selecione o certificado para continuar")
        root.iconify()
        WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, '//*[@id="sairAplicacao"]')))

        workbook = openpyxl.load_workbook(caminho_planilha_var.get())
        sheet_empresas = workbook.active        
        documento = sheet_empresas.cell(row=int(linha_ini.get()), column=3).value        
        empresas = list(sheet_empresas.iter_rows(min_row=int(linha_ini.get()), max_row=int(linha_fim.get())))

        for linha in sheet_empresas.iter_rows(min_row=int(linha_ini.get()), max_row=int(linha_fim.get())):
            pasta_empresa = criar_pasta(f'{linha[0].value} - {linha[1].value.rstrip()}')
            WebDriverWait(driver, 120).until(EC.presence_of_element_located((By.XPATH, '//*[@id="menuDownload"]')))
            driver.get("https://www.esocial.gov.br/portal/download/Pedido/Consulta")
            WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="conteudo-pagina"]/form/section/div/div[4]/input')))
            driver.find_element(By.XPATH, '//*[@id="conteudo-pagina"]/form/section/div/div[4]/input').click()
            driver.execute_script(script_download)
            print("Baixando arquivos...")                        
            WebDriverWait(driver, 120).until(EC.element_to_be_clickable((By.XPATH, '/html/body/div[13]/button[2]')))
            driver.find_element(By.XPATH, '/html/body/div[13]/button[2]').click()
            # Função para mover arquivos .zip
            def verificar_e_mover_arquivos_zip():
                arquivos_baixados = os.listdir(download_dir)
                for arquivo in arquivos_baixados:
                    if arquivo.endswith(".zip"):  # Verifica se o arquivo é um ZIP
                        caminho_arquivo = os.path.join(download_dir, arquivo)
                        shutil.move(caminho_arquivo, os.path.join(pasta_empresa, arquivo))
                        print(f"Arquivo {arquivo} movido para {pasta_empresa}")
            
            intervalo_verificacao = 20  # segundos
            time.sleep(10)
            download_dir = caminho_pasta_salvar_var.get().rstrip().replace('/', '\\')            
            loop = True            
            while loop == True:
                arquivos_baixados = os.listdir(download_dir)
                existe_zip = any(arquivo.endswith('.zip') for arquivo in arquivos_baixados)
                if existe_zip:
                    verificar_e_mover_arquivos_zip()
                    print(f"Aguardando {intervalo_verificacao} segundos para próxima verificação de arquivos baixados...")
                    time.sleep(intervalo_verificacao)                    
                else:
                    loop = False
                    print(f"Todos arquivos da empresa {cnpj} foram baixados!")    
        
        print("Baixa de arquivos finalizada!")            
        driver.quit()

def meses_buscar():
    # Função que será chamada quando o botão for clicado
    meses_buscar_var.get()     

def selecionar_arquivo():
    arquivo = filedialog.askopenfilename(filetypes=[("Arquivos XLSX", "*.xlsx")])
    if arquivo:
        caminho_planilha_var.set(arquivo)

def selecionar_pasta_salvar():
    pasta_salvar = filedialog.askdirectory()
    if pasta_salvar:
        caminho_pasta_salvar_var.set(pasta_salvar)

# Criando a janela principal
root = tk.Tk()
root.title("Download eSocial")
root.resizable(False, False)

# Labels
labels = [
    "Caminho da planilha:",
    "Linha inicial da planilha:",
    "Linha final da planilha:",
    "Meses a buscar:", 
    "Solicitar (1) / Baixar (2):",   
    "Salvar arquivos:",    
    "Certificado próprio:"
]

# Variáveis para armazenar valores
caminho_planilha_var = tk.StringVar()
caminho_pasta_salvar_var = tk.StringVar()
certificado_proprio_var = tk.BooleanVar()
linha_ini = tk.IntVar()
linha_fim = tk.IntVar()

# Posicionamento dos labels e entradas
for i, label_text in enumerate(labels):
    label = tk.Label(root, text=label_text)
    label.grid(row=i, column=0, padx=5, pady=5, sticky="w")
    
    if i == 0:
        entry = tk.Entry(root, textvariable=caminho_planilha_var, width=40)
        entry.grid(row=i, column=1, padx=5, pady=5, sticky="ew")
        button = tk.Button(root, text="Selecionar", command=selecionar_arquivo)
        button.grid(row=i, column=2, padx=5, pady=5)
    elif i == 1:
        entry = tk.Entry(root, textvariable=linha_ini)
        entry.grid(row=i, column=1, padx=5, pady=5, sticky="ew")
    elif i == 2:
        entry = tk.Entry(root, textvariable=linha_fim)
        entry.grid(row=i, column=1, padx=5, pady=5, sticky="ew") 
    elif i == 3:
        meses_buscar_var = tk.IntVar()
        dropdown = ttk.Combobox(root, values=[1, 2, 3, 4, 5, 6], textvariable=meses_buscar_var, state="readonly")
        dropdown.grid(row=i, column=1, padx=5, pady=5, sticky="ew")
    elif i == 4:
        solicitar_baixar_var = tk.IntVar()
        dropdown = ttk.Combobox(root, values=[1, 2], textvariable=solicitar_baixar_var, state="readonly")
        dropdown.grid(row=i, column=1, padx=5, pady=5, sticky="ew")
    elif i == 5:
        entry = tk.Entry(root, textvariable=caminho_pasta_salvar_var)
        entry.grid(row=i, column=1, padx=5, pady=5, sticky="ew")
        button = tk.Button(root, text="Selecionar", command=selecionar_pasta_salvar)
        button.grid(row=i, column=2, padx=5, pady=5)
    elif i == 6:
        checkbutton = tk.Checkbutton(root, variable=certificado_proprio_var)
        checkbutton.grid(row=i, column=1, padx=5, pady=5, sticky="w")
    else:
        entry = tk.Entry(root)
        entry.grid(row=i, column=1, padx=5, pady=5, sticky="ew")

# Botão
button = tk.Button(root, text="Iniciar", command=solicitar_ou_baixar, width=10)
button2 = tk.Button(root, text="Cancelar", command=root.quit, width=10)
button.grid(row=len(labels), column=0, pady=5, padx=5)
button2.grid(row=len(labels), column=1, pady=5, padx=5)

root.mainloop()